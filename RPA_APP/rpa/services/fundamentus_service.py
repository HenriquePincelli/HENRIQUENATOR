from RPA_APP.rpa.services.service import Service
from botcity.web import By
from time import sleep
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment


class FundamentusService(Service):

    # >>>>>>>>>Function to extract all data from Aljazeera website>>>>>>>>>
    def extract_fundamentus_data(self, bot):
        try:
            # Give time to load all elements in HTML
            sleep(5)

            # Extract papers data
            papers = bot.find_element("resultado", By.ID, waiting_time=10000).text.split("\n")
            # papers = papers.split("\n")
            papers.pop(0)
            # <<<<<<<<<Extract papers data<<<<<<<<<

            return {
                "status": True,
                "bot": bot,
                "data": papers
            }
        except Exception as e:
            # Tracing the Error
            return self.trace_error(e, bot)
    # <<<<<<<<<Function to extract all data from Aljazeera website<<<<<<<<<

    # >>>>>>>>>Make a excel file with fundamentus data>>>>>>>>>
    def make_fundamentus_excel_file(self, data, directory, excel_filename):
        try:
            # Columns names
            columns_list = ["Papel", "P/L", "P/VP", "Div.Yield", "ROE", "EV/EBITDA", "Cresc. Rec.5a"]

            # Creating a new workbook
            wb = Workbook()
            # Selecting an active sheet
            ws = wb.active
            # Adding column names to the first row
            ws.append(columns_list)

            # >>>>>>>>>Iterating over the data list and formatting the values>>>>>>>>>
            for v in data:
                value = ' '.join(v.split()).split(" ")
                # Ensuring the length matches the expected format
                if len(value) == 21:
                    # Parsing and appending row data to the sheet
                    row_list = [str(value[0]), float(value[2].replace(".", "").replace(",", ".").replace("%", "")), 
                                float(value[3].replace(".", "").replace(",", ".").replace("%", "")), 
                                float(value[5].replace(".", "").replace(",", ".").replace("%", "")), 
                                float(value[11].replace(".", "").replace(",", ".").replace("%", "")), 
                                float(value[16].replace(".", "").replace(",", ".").replace("%", "")), 
                                float(value[20].replace(".", "").replace(",", ".").replace("%", ""))]
                    ws.append(row_list)
            # <<<<<<<<<Iterating over the data list and formatting the values<<<<<<<<<

            # Setting a thin border style for all cells
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                top=Side(style='thin'), bottom=Side(style='thin'))

            # >>>>>>>>>Auto-adjusting column widths based on the longest cell content>>>>>>>>>
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 4)
                ws.column_dimensions[column].width = adjusted_width
            # <<<<<<<<<Auto-adjusting column widths based on the longest cell content<<<<<<<<<

            # >>>>>>>>>Styling the header row with bold font and black background>>>>>>>>>
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            # <<<<<<<<<Styling the header row with bold font and black background<<<<<<<<<

            # >>>>>>>>>Applying styles to all rows and cells>>>>>>>>>
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(columns_list)):
                for cell in row:
                    # Applying border and center alignment to all cells
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                    # Making the first column bold
                    if cell.column == 1:
                        cell.font = Font(bold=True)

                    # Applying the header styles to the first row
                    if cell.row == 1:
                        cell.font = header_font
                        cell.fill = header_fill
            # <<<<<<<<<Applying styles to all rows and cells<<<<<<<<<

            # Saving the Excel file in the specified directory with the given filename
            wb.save(f"{directory}{excel_filename}")
            return {"status": True}
        except Exception as e:
            # Tracing the Error
            return self.trace_error(e)
    # <<<<<<<<<Make a excel file with fundamentus data<<<<<<<<<
